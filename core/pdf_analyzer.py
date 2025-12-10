# core/pdf_analyzer.py
"""
PDF 파서 + 계산 규칙을 이용해
성명별 집계를 만들고, 지급조서 템플릿(현재 구조 고정)에 결과를 채워넣는 모듈.

템플릿 전제:
- C4: '성명' 헤더
- C5부터 데이터 입력 (연번/직급/성명 등은 비어있어도 상관없음)
- 기본 21명(= C5~C25)까지 들어가는 양식이고,
  이 경우 합계는 원래 H26 셀에 있어야 한다.
- 21명을 초과하면 H26 위에 줄을 추가로 삽입해서,
  마지막 사람 바로 아래 H열에 합계를 둔다.
- 인원수 제한은 없으며, 22명 이상일 때 필요한 만큼 줄을 삽입해 30명, 40명 등
  대규모 명단도 합계 행이 자동으로 아래로 이동한다.
"""

from __future__ import annotations
from typing import Tuple
import io

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from .pdf_parser import parse_pdf_to_rows
from .rules import compute_allowance_by_person


# 템플릿 고정 위치
HEADER_ROW = 4     # C4 = 성명 헤더
DATA_START_ROW = 5 # C5부터 데이터
COL_NO = 1         # A열 연번
COL_NAME = 3       # C열 성명
COL_UNDER4 = 4     # D열 4시간 미만
COL_OVER4 = 5      # E열 4시간 이상
COL_CAR = 6        # F열 차량사용횟수
COL_PDF_AMT = 8    # H열 PDF 기준 지급액(원)
COL_CALC_AMT = 12  # L열: 규칙상 올바른 금액 (PDF 금액과 다를 때만 표시)


# ─────────────────────────────────────
# 1. PDF 행 단위 → 사람별 집계
# ─────────────────────────────────────

def summarize_pdf_by_person(df_rows: pd.DataFrame) -> pd.DataFrame:
    """
    행단위 df_rows (parse_pdf_to_rows 결과)를 받아
    PDF 기준 성명별 집계 DataFrame 생성.

    기대 열:
      - '성명'
      - 'minutes'   : 해당 출장의 총 출장시간(분)
      - 'car_used'  : 공용차량 사용 여부 (True/False 또는 0/1)
      - 'amount_pdf': PDF에 찍힌 해당 출장의 금액(원)

    반환:
        index: 성명
        columns:
          - 총지급액_숫자
          - 4시간미만
          - 4시간이상
          - 차량사용횟수
    """
    df = df_rows.copy()

    # 안전장치 (열 없으면 0으로)
    for col in ["minutes", "car_used", "amount_pdf"]:
        if col not in df.columns:
            df[col] = 0

    df["minutes"] = pd.to_numeric(df["minutes"], errors="coerce").fillna(0).astype(int)
    df["car_used"] = df["car_used"].astype(bool)
    df["amount_pdf"] = pd.to_numeric(df["amount_pdf"], errors="coerce").fillna(0).astype(int)

    df["is_under4"] = (df["minutes"] > 0) & (df["minutes"] < 240)
    df["is_over4"] = df["minutes"] >= 240
    df["car_cnt"] = df["car_used"].astype(int)

    grouped = df.groupby("성명", dropna=True).agg(
        총지급액_숫자=("amount_pdf", "sum"),
        _4시간미만=("is_under4", "sum"),
        _4시간이상=("is_over4", "sum"),
        차량사용횟수=("car_cnt", "sum"),
    )

    grouped = grouped.rename(columns={"_4시간미만": "4시간미만", "_4시간이상": "4시간이상"})
    for c in ["총지급액_숫자", "4시간미만", "4시간이상", "차량사용횟수"]:
        grouped[c] = grouped[c].fillna(0).astype(int)

    return grouped


# ─────────────────────────────────────
# 2. 템플릿에 집계 결과 채워넣기
# ─────────────────────────────────────

def fill_template_with_summary(template_bytes: bytes, summary: pd.DataFrame) -> bytes:
    """
    지급조서 템플릿(엑셀 바이너리)에 summary 정보를 채워넣고,
    결과 엑셀을 bytes로 반환.

    summary:
      index = 성명
      columns 에 최소한 아래가 포함되도록 기대:
        - 총지급액_숫자
        - 4시간미만
        - 4시간이상
        - 차량사용횟수
        - 계산_총지급액
        - 차이
    """
    summary = summary.copy()
    summary = summary.fillna(0)

    for col in ["총지급액_숫자", "4시간미만", "4시간이상", "차량사용횟수", "계산_총지급액", "차이"]:
        if col in summary.columns:
            summary[col] = pd.to_numeric(summary[col], errors="coerce").fillna(0).astype(int)

    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    red_bold_font = Font(color="FF0000", bold=True)

    # 성명 목록 (보기 좋게 정렬)
    people = list(summary.sort_index().iterrows())
    n_people = len(people)

    # 사람 없으면 템플릿만 그대로 반환
    if n_people == 0:
        out_empty = io.BytesIO()
        wb.save(out_empty)
        out_empty.seek(0)
        return out_empty.getvalue()

    # ────────── 합계 위치 설계 ──────────
    # 기본: 21명(=C5~C25), 합계는 H26 (BASE_SUM_ROW)
    MAX_ROWS_WITHOUT_INSERT = 21
    BASE_SUM_ROW = DATA_START_ROW + MAX_ROWS_WITHOUT_INSERT  # 5 + 21 = 26

    # 인원 제한 없이 21명을 넘으면 필요한 만큼 합계 행 위에 줄 삽입
    if n_people > MAX_ROWS_WITHOUT_INSERT:
        extra_rows = n_people - MAX_ROWS_WITHOUT_INSERT
        ws.insert_rows(BASE_SUM_ROW, amount=extra_rows)
        total_row = BASE_SUM_ROW + extra_rows  # 합계 행은 밑으로 밀림
    else:
        total_row = BASE_SUM_ROW  # 21명 이하이면 H26 고정

    # ────────── 데이터 채우기 ──────────
    current_row = DATA_START_ROW
    for idx, (name, row) in enumerate(people, start=1):
        name_str = str(name).strip()
        if not name_str:
            continue

        under4 = int(row.get("4시간미만", 0) or 0)
        over4 = int(row.get("4시간이상", 0) or 0)
        car_cnt = int(row.get("차량사용횟수", 0) or 0)
        amount_pdf = int(row.get("총지급액_숫자", 0) or 0)
        amount_calc = int(row.get("계산_총지급액", 0) or 0)

        # 연번(A열)
        ws.cell(row=current_row, column=COL_NO).value = idx

        # 성명(C열)
        ws.cell(row=current_row, column=COL_NAME).value = name_str

        # 4시간 미만/이상/차량 (0이면 공란)
        ws.cell(row=current_row, column=COL_UNDER4).value = "" if under4 == 0 else under4
        ws.cell(row=current_row, column=COL_OVER4).value = "" if over4 == 0 else over4
        ws.cell(row=current_row, column=COL_CAR).value = "" if car_cnt == 0 else car_cnt

        # PDF 금액(H열)
        cell_h = ws.cell(row=current_row, column=COL_PDF_AMT)
        cell_h.value = "" if amount_pdf == 0 else amount_pdf


        
       codex/fix-layout-alignment-issue-j8zv8a
        # 계산 금액(L열) - '차이' 기준으로 표시 (계산액이 0이어도 차이가 있으면 표기)
        diff = int(row.get("차이", amount_calc - amount_pdf) or 0)
        cell_l = ws.cell(row=current_row, column=COL_CALC_AMT)
        if diff != 0:
        # 계산 금액(L열) - PDF와 다를 때 표시 (계산액이 0이어도 차이가 있으면 표기)
        cell_l = ws.cell(row=current_row, column=COL_CALC_AMT)
        if amount_pdf != amount_calc:
        main
            cell_l.value = amount_calc
            cell_h.font = red_bold_font
            cell_l.font = red_bold_font
        else:
            cell_l.value = ""

        current_row += 1

    last_data_row = current_row - 1  # 마지막 사람 행

    # ────────── 합계 수식(H열) 설정 ──────────
    # 템플릿에서 A26:G26은 합쳐진 셀("합계" 글자 있음)이라
    # G열에는 글자를 쓰지 않고, H열에만 수식 넣는다.
    if last_data_row >= DATA_START_ROW:
        first_cell = ws.cell(row=DATA_START_ROW, column=COL_PDF_AMT).coordinate
        last_cell = ws.cell(row=last_data_row, column=COL_PDF_AMT).coordinate
        ws.cell(row=total_row, column=COL_PDF_AMT).value = f"=SUM({first_cell}:{last_cell})"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ─────────────────────────────────────
# 3. 앱에서 부르는 메인 함수
# ─────────────────────────────────────

def analyze_pdf_and_template(
    pdf_bytes: bytes,
    template_bytes: bytes,
) -> Tuple[pd.DataFrame, bytes]:
    """
    PDF + 템플릿을 받아서:
      - 성명별 요약 DataFrame(summary)
      - 템플릿에 결과 채운 엑셀 bytes
    를 반환.
    """
    # 1) PDF 파싱 (행 단위 데이터)
    df_rows = parse_pdf_to_rows(pdf_bytes)

    # 2) PDF 기준 집계
    pdf_summary = summarize_pdf_by_person(df_rows)

    # 3) 규칙에 따른 실제 지급액 계산 (rules.py)
    calc_totals = compute_allowance_by_person(df_rows)
    # calc_totals: index=성명, value=계산된 총지급액

    # 4) summary 결합
    summary = pdf_summary.copy()
    # calc_totals를 summary index에 맞춰 align
    summary["계산_총지급액"] = calc_totals.reindex(summary.index).fillna(0).astype(int)
    summary["차이"] = summary["계산_총지급액"] - summary["총지급액_숫자"]

    # 5) 템플릿 채우기
    result_bytes = fill_template_with_summary(template_bytes, summary)

    return summary, result_bytes